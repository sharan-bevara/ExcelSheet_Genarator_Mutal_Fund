"""
NGEN Markets – Equity Fund Rankings Generator
Processes ALL 2,564 equity-related funds:
  - Equity Scheme - * (all 12 sub-categories)  →  2,392 funds
  - Hybrid Scheme - Equity Savings              →    172 funds
All plan types included (Regular/Direct/Growth/IDCW/Other)
52 batches · ~10 min · ~$0.18 on gpt-4o-mini
"""

import streamlit as st
import pandas as pd
import json, time, math, re
from datetime import datetime
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except ImportError:
    st.error("openai package not found.  Run:  pip install openai")
    st.stop()

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NGEN Markets – Equity Rankings",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=Syne:wght@700;800&display=swap');
html,body,[class*="css"]{font-family:'Space Grotesk',sans-serif;}
.stApp{background:#0a0e1a;color:#e8eaf6;}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#0d1226,#111827);border-right:1px solid #1e2a45;}
[data-testid="stSidebar"] *{color:#c9d1e8 !important;}

.hero{background:linear-gradient(135deg,#0d1b3e,#102040,#0a1628);border:1px solid #1e3a5f;border-radius:16px;padding:2rem 2.5rem;margin-bottom:1.5rem;}
.hero-badge{display:inline-block;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.35);color:#60a5fa;font-size:.72rem;font-weight:600;letter-spacing:.08em;text-transform:uppercase;padding:3px 10px;border-radius:20px;margin-bottom:.75rem;}
.hero-title{font-family:'Syne',sans-serif;font-size:2rem;font-weight:800;color:#fff;letter-spacing:-.03em;margin:0 0 .25rem;}
.hero-sub{font-size:.95rem;color:#7b91b8;margin:0;}

.section-hdr{font-family:'Syne',sans-serif;font-size:1.1rem;font-weight:700;color:#e2e8f0;border-bottom:1px solid #1e2a45;padding-bottom:.5rem;margin-bottom:1rem;}

.metric-row{display:flex;gap:1rem;margin-bottom:1.5rem;flex-wrap:wrap;}
.metric-card{background:#111827;border:1px solid #1e2a45;border-radius:12px;padding:1rem 1.5rem;flex:1;min-width:130px;position:relative;overflow:hidden;}
.metric-card::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#3b82f6,#8b5cf6);}
.metric-label{font-size:.72rem;color:#4b6080;font-weight:600;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.3rem;}
.metric-value{font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:700;color:#fff;line-height:1;}
.metric-sub{font-size:.75rem;color:#4b6080;margin-top:.2rem;}

.filter-box{background:rgba(52,211,153,.06);border:1px solid rgba(52,211,153,.25);border-radius:10px;padding:1rem 1.25rem;margin-bottom:1rem;font-size:.875rem;color:#6ee7b7;line-height:1.7;}
.info-box{background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.25);border-radius:10px;padding:1rem 1.25rem;margin-bottom:1rem;font-size:.875rem;color:#93b4d8;line-height:1.6;}
.warn-box{background:rgba(251,191,36,.06);border:1px solid rgba(251,191,36,.25);border-radius:10px;padding:1rem 1.25rem;margin-bottom:1rem;font-size:.875rem;color:#fcd34d;line-height:1.6;}

.cat-grid{display:flex;flex-wrap:wrap;gap:.5rem;margin:.75rem 0;}
.cat-pill{background:rgba(59,130,246,.1);border:1px solid rgba(59,130,246,.3);color:#93c5fd;font-size:.75rem;font-weight:600;padding:4px 12px;border-radius:20px;}
.cat-pill span{background:rgba(59,130,246,.25);border-radius:10px;padding:1px 6px;margin-left:4px;font-size:.7rem;}
.cat-pill.hybrid{background:rgba(167,139,250,.1);border:1px solid rgba(167,139,250,.3);color:#c4b5fd;}
.cat-pill.hybrid span{background:rgba(167,139,250,.25);}

.plan-grid{display:flex;flex-wrap:wrap;gap:.5rem;margin:.5rem 0;}
.plan-pill{font-size:.72rem;font-weight:600;padding:3px 10px;border-radius:20px;}
.plan-rg{background:rgba(52,211,153,.12);border:1px solid rgba(52,211,153,.3);color:#34d399;}
.plan-dg{background:rgba(96,165,250,.12);border:1px solid rgba(96,165,250,.3);color:#60a5fa;}
.plan-ri{background:rgba(251,191,36,.1);border:1px solid rgba(251,191,36,.3);color:#fcd34d;}
.plan-di{background:rgba(167,139,250,.12);border:1px solid rgba(167,139,250,.3);color:#a78bfa;}
.plan-ot{background:rgba(156,163,175,.1);border:1px solid rgba(156,163,175,.3);color:#9ca3af;}

.progress-card{background:#111827;border:1px solid #1e2a45;border-radius:12px;padding:1.25rem 1.5rem;margin-bottom:1rem;}
.log-box{background:#060a14;border:1px solid #1e2a45;border-radius:8px;padding:1rem;font-family:'Courier New',monospace;font-size:.78rem;color:#4ade80;max-height:260px;overflow-y:auto;white-space:pre-wrap;line-height:1.5;}
.badge-run{color:#60a5fa;font-weight:600;}
.badge-err{color:#f87171;font-weight:600;}

.formula-box{background:linear-gradient(135deg,#0d1b3e,#0a1628);border:1px solid #1e3a5f;border-radius:10px;padding:1rem 1.25rem;font-size:.82rem;color:#93b4d8;line-height:1.7;}
.formula-box strong{color:#60a5fa;}
.wpill{font-size:.72rem;padding:3px 10px;border-radius:20px;font-weight:600;display:inline-block;margin:2px;}
.wh{background:rgba(52,211,153,.12);border:1px solid rgba(52,211,153,.3);color:#34d399;}
.wl{background:rgba(248,113,113,.12);border:1px solid rgba(248,113,113,.3);color:#f87171;}

.stButton>button{background:linear-gradient(135deg,#1d4ed8,#2563eb);color:white;border:none;border-radius:8px;font-family:'Space Grotesk',sans-serif;font-weight:600;padding:.6rem 1.5rem;font-size:.9rem;width:100%;transition:all .2s;}
.stButton>button:hover{background:linear-gradient(135deg,#2563eb,#3b82f6);transform:translateY(-1px);box-shadow:0 4px 15px rgba(59,130,246,.3);}
.stButton>button:disabled{background:#1e2a45 !important;color:#4b6080 !important;}
.stTextInput input{background:#111827 !important;border:1px solid #1e2a45 !important;color:#e8eaf6 !important;border-radius:8px !important;}
label{color:#7b91b8 !important;font-size:.85rem !important;}
.stProgress>div>div{background:linear-gradient(90deg,#3b82f6,#8b5cf6) !important;}
h1,h2,h3{color:#e2e8f0 !important;font-family:'Syne',sans-serif !important;}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
# Hard-coded fund filter — exactly the 2 category groups wanted
TARGET_CATEGORIES = {
    "equity":  lambda cat: str(cat).startswith("Equity Scheme"),
    "hybrid":  lambda cat: cat == "Hybrid Scheme - Equity Savings",
}

COLUMNS = [
    "Fund Name","AUM Cr.","TER","PE","PB",
    "Top 3 Holdings","Top 5 Holdings","Top 10 Holdings","Top 20 Holdings",
    "Sharpe","Sortino","St Dev","Inception","Age in Yrs",
]

WEIGHTS = {
    "AUM Cr.":          ("higher", 9),
    "TER":              ("lower",  8),
    "PE":               ("lower",  7),
    "PB":               ("lower",  7),
    "Top 3 Holdings":   ("lower",  6),
    "Top 5 Holdings":   ("lower",  6),
    "Top 10 Holdings":  ("lower",  6),
    "Top 20 Holdings":  ("lower",  6),
    "Sharpe":           ("higher", 15),
    "Sortino":          ("higher", 13),
    "St Dev":           ("lower",  9),
    "Inception":        ("higher", 3),
    "Age in Yrs":       ("higher", 5),
}

SYSTEM_PROMPT = """You are a financial data expert on Indian mutual funds.
Return ONLY a valid JSON array — no markdown, no code fences, no explanation.
Each element must have exactly these keys:
  "Fund Name","AUM Cr.","TER","PE","PB",
  "Top 3 Holdings","Top 5 Holdings","Top 10 Holdings","Top 20 Holdings",
  "Sharpe","Sortino","St Dev","Inception","Age in Yrs"

Rules (all values must be plain numbers, NOT strings):
- AUM Cr.        : AUM in Indian Crores              e.g. 45230
- TER            : Expense ratio %                   e.g. 1.85   (Direct plans are ~0.5-1.0 lower)
- PE             : Wtd avg portfolio Price/Earnings   e.g. 28.4
- PB             : Wtd avg portfolio Price/Book       e.g. 4.1
- Top 3/5/10/20  : % weight in top N holdings         e.g. 18.4
- Sharpe         : 3-yr Sharpe ratio                  e.g. 0.82
- Sortino        : 3-yr Sortino ratio                 e.g. 1.04
- St Dev         : 3-yr annualised std deviation %    e.g. 14.9
- Inception      : CAGR since inception %             e.g. 16.4
- Age in Yrs     : Fund age in years (1 decimal)      e.g. 29.0
- For Hybrid Equity Savings funds: PE/PB/Holdings reflect equity portion only
- Direct plans share same PE/PB/Holdings/Sharpe/Sortino/StDev as Regular plan of same scheme
- Use null for any value you cannot confirm
- Return ONLY the JSON array, nothing else"""

# ── Helpers ────────────────────────────────────────────────────────────────────
def load_funds(uploaded_file):
    """Load CSV and hard-filter to all 2,564 target funds."""
    df = pd.read_csv(uploaded_file)
    total_csv = len(df)
    mask = (
        df["Scheme Category"].str.startswith("Equity Scheme", na=False) |
        (df["Scheme Category"] == "Hybrid Scheme - Equity Savings")
    )
    funds = df[mask].copy().reset_index(drop=True)
    return funds, total_csv

def get_plan_type(name):
    n = str(name).lower()
    is_direct = "direct" in n
    is_growth = "growth" in n
    is_idcw   = any(x in n for x in ["idcw", "dividend"])
    if   is_direct and is_growth: return "Direct - Growth"
    elif is_direct and is_idcw:   return "Direct - IDCW"
    elif is_direct:               return "Direct - Other"
    elif is_growth:               return "Regular - Growth"
    elif is_idcw:                 return "Regular - IDCW"
    else:                         return "Regular - Other"

def clean_num(val):
    if val is None: return None
    if isinstance(val, float) and math.isnan(val): return None
    if isinstance(val, str):
        val = val.replace("%", "").replace(",", "").strip()
        try:    return float(val)
        except: return None
    try:    return float(val)
    except: return None

def calculate_score(row):
    score = 0.0
    for col, (direction, weight) in WEIGHTS.items():
        val = clean_num(row.get(col))
        if val is None: continue
        score += val * weight if direction == "higher" else -(val * weight)
    return round(score, 4)

def build_excel_export(df):
    """Build Excel export exactly matching the sample format with 5-row header block."""
    export_cols = [
        "Fund Name","AUM Cr.","TER","PE","PB",
        "Top 3 Holdings","Top 5 Holdings","Top 10 Holdings","Top 20 Holdings",
        "Sharpe","Sortino","St Dev","Inception","Age in Yrs","score","Rank"
    ]
    directions = ["Higher/lower","higher","lower","lower","lower","lower","lower","lower","lower","higher","higher","lower","higher","higher","",""]
    weights    = ["Weightage","9","8","7","7","6","6","6","6","15","13","9","3","5","score","Rank"]

    # Colors
    dark_bg    = "0A0E1A"; header_bg  = "0D1226"; col_hdr_bg = "111827"
    higher_bg  = "0D2818"; lower_bg   = "2D0A0A"; weight_bg  = "0D1B3E"
    data_bg1   = "0F1629"; data_bg2   = "111827"; blue_accent= "3B82F6"
    green_col  = "34D399"; red_col    = "F87171"; white      = "FFFFFF"
    gold       = "FCD34D"; light_gray = "C9D1E8"; muted      = "7B91B8"

    higher_set = {"AUM Cr.","Sharpe","Sortino","Inception","Age in Yrs"}
    lower_set  = {"TER","PE","PB","Top 3 Holdings","Top 5 Holdings","Top 10 Holdings","Top 20 Holdings","St Dev"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Rankings"
    ws.sheet_properties.tabColor = blue_accent

    n = len(export_cols)

    # Row 1: Title
    ws.merge_cells(f"A1:{get_column_letter(n)}1")
    c = ws["A1"]
    c.value = "NGEN Markets Filter Results"
    c.font = Font(name="Arial", bold=True, size=14, color=white)
    c.fill = PatternFill("solid", fgColor=dark_bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Blank + Notes
    ws.merge_cells(f"A2:{get_column_letter(n-1)}2")
    ws["A2"].fill = PatternFill("solid", fgColor=header_bg)
    nc = ws[f"{get_column_letter(n)}2"]
    nc.value = "Notes"
    nc.font = Font(name="Arial", italic=True, color=muted)
    nc.fill = PatternFill("solid", fgColor=header_bg)
    nc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Row 3: Column headers
    for i, col in enumerate(export_cols, 1):
        c = ws.cell(row=3, column=i, value=col)
        c.font = Font(name="Arial", bold=True, size=10, color=white)
        c.fill = PatternFill("solid", fgColor=col_hdr_bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=blue_accent))
    ws.row_dimensions[3].height = 32

    # Row 4: Higher/Lower
    notes_r4 = "Whether the parameter should have a higher value or a lower value for a positive outcome"
    for i, val in enumerate(directions, 1):
        c = ws.cell(row=4, column=i, value=val)
        if val == "higher":
            c.font = Font(name="Arial", bold=True, size=9, color=green_col)
            c.fill = PatternFill("solid", fgColor=higher_bg)
        elif val == "lower":
            c.font = Font(name="Arial", bold=True, size=9, color=red_col)
            c.fill = PatternFill("solid", fgColor=lower_bg)
        else:
            c.font = Font(name="Arial", bold=True, size=9, color=light_gray)
            c.fill = PatternFill("solid", fgColor=header_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=n+1, value=notes_r4).font = Font(name="Arial", italic=True, size=8, color=muted)
    ws.row_dimensions[4].height = 18

    # Row 5: Weightage
    notes_r5  = "Weightages given to the parameters adding up to 100"
    formula_n = "(The weightage × corresponding value for all 'higher' columns) − (The weightage × corresponding values for all 'lower' columns) = score"
    for i, val in enumerate(weights, 1):
        c = ws.cell(row=5, column=i, value=val)
        c.font = Font(name="Arial", bold=True, size=9, color=gold)
        c.fill = PatternFill("solid", fgColor=weight_bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=n+1, value=notes_r5).font = Font(name="Arial", italic=True, size=8, color=muted)
    ws.cell(row=5, column=n+2, value=formula_n).font = Font(name="Arial", italic=True, size=8, color=muted)
    ws.row_dimensions[5].height = 18

    # Prepare data
    out = df.copy()
    out["score"] = out["Score"]
    for col in export_cols:
        if col not in out.columns:
            out[col] = None

    # Data rows
    for r_idx, (_, row) in enumerate(out.iterrows(), 6):
        bg = data_bg1 if r_idx % 2 == 0 else data_bg2
        for c_idx, col in enumerate(export_cols, 1):
            val = row.get(col)
            if isinstance(val, float) and math.isnan(val):
                val = None
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            if c_idx == 1:
                c.font = Font(name="Arial", size=9, color=light_gray)
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx == n:  # Rank
                c.font = Font(name="Arial", bold=True, size=9, color=gold)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = "0"
            elif c_idx == n-1:  # Score
                c.font = Font(name="Arial", size=9, color=white)
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0.0000"
            elif col in higher_set:
                c.font = Font(name="Arial", size=9, color=green_col)
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif col in lower_set:
                c.font = Font(name="Arial", size=9, color=red_col)
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.font = Font(name="Arial", size=9, color=light_gray)
                c.alignment = Alignment(horizontal="right", vertical="center")
            # Number formats
            fmt_map = {
                "AUM Cr.": "#,##0", "TER": "0.00", "PE": "0.00", "PB": "0.00",
                "Top 3 Holdings": "0.00", "Top 5 Holdings": "0.00",
                "Top 10 Holdings": "0.00", "Top 20 Holdings": "0.00",
                "Sharpe": "0.000", "Sortino": "0.000",
                "St Dev": "0.00", "Inception": "0.00", "Age in Yrs": "0.0",
            }
            if col in fmt_map:
                c.number_format = fmt_map[col]
        ws.row_dimensions[r_idx].height = 16

    # Column widths
    widths = [52,12,7,8,7,14,14,15,15,9,9,9,10,10,16,7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[get_column_letter(n+1)].width = 55
    ws.column_dimensions[get_column_letter(n+2)].width = 80

    ws.freeze_panes = "B6"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def build_prompt(fund_names):
    numbered = "\n".join(f"{i+1}. {n}" for i, n in enumerate(fund_names))
    return (
        f"Provide financial metrics for these {len(fund_names)} Indian mutual fund NAV plans "
        f"(equity and equity-savings hybrid funds).\n"
        f"Source: AMFI, Value Research, Morningstar India.\n\n"
        f"{numbered}\n\n"
        f"Return a JSON array with exactly {len(fund_names)} objects in the same order as listed."
    )

def call_openai(client, fund_names, model, retries=3):
    for attempt in range(retries):
        try:
            kwargs = dict(
                model=model,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": build_prompt(fund_names)},
                ],
                temperature=0.1,
                max_tokens=4096,
            )
            if model in ("gpt-4o", "gpt-4o-mini", "gpt-4-turbo"):
                kwargs["response_format"] = {"type": "json_object"}
            resp   = client.chat.completions.create(**kwargs)
            raw    = resp.choices[0].message.content.strip()
            raw    = re.sub(r"^```[a-z]*\n?", "", raw)
            raw    = re.sub(r"\n?```$",        "", raw)
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                for v in parsed.values():
                    if isinstance(v, list):
                        parsed = v
                        break
            if isinstance(parsed, list):
                return parsed
            return []
        except Exception:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return []

def mcard(label, value, sub=""):
    return (
        f'<div class="metric-card"><div class="metric-label">{label}</div>'
        f'<div class="metric-value">{value}</div>'
        + (f'<div class="metric-sub">{sub}</div>' if sub else "")
        + '</div>'
    )

# ── Session state ──────────────────────────────────────────────────────────────
DEFAULTS = {
    "running": False, "results": [], "logs": [],
    "processed": 0, "errors": 0, "total": 0,
    "final_df": None, "start_time": None,
    "funds_df": None, "total_csv": 0,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        "<div style='font-family:Syne,sans-serif;font-size:1.15rem;"
        "font-weight:800;color:#fff;margin-bottom:1.5rem;'>⚙️ Configuration</div>",
        unsafe_allow_html=True)

    st.markdown("**OpenAI API Key**")
    api_key = st.text_input("API Key", type="password",
                             placeholder="sk-...", label_visibility="collapsed")

    st.markdown("**Model**")
    model = st.selectbox("Model",
        ["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"],
        label_visibility="collapsed")

    st.markdown("**Batch size** (funds per API call)")
    batch_size = st.slider("Batch", 10, 50, 50, 5, label_visibility="collapsed")

    st.markdown("**Delay between batches (sec)**")
    delay = st.slider("Delay", 0.5, 5.0, 1.0, 0.5, label_visibility="collapsed")

    st.markdown("**Max batches to run** (0 = all)")
    max_batches = st.slider("Max Batches", 0, 52, 0, 1, label_visibility="collapsed")
    if max_batches > 0:
        st.markdown(f"<div class='warn-box'>⚠️ Will process <strong>{max_batches} batch{'es' if max_batches>1 else ''}</strong> → ~<strong>{max_batches * 50:,} funds</strong></div>", unsafe_allow_html=True)

    st.divider()
    st.markdown("""<div class='formula-box'><strong>📐 Score Formula</strong><br><br>
<span style='color:#34d399'>Σ(weight × value)</span> — higher is better<br>
<span style='color:#f87171'>− Σ(weight × value)</span> — lower is better<br><br>
<span class='wpill wh'>AUM ×9</span><span class='wpill wh'>Sharpe ×15</span>
<span class='wpill wh'>Sortino ×13</span><span class='wpill wh'>Inception ×3</span>
<span class='wpill wh'>Age ×5</span><br>
<span class='wpill wl'>TER ×8</span><span class='wpill wl'>PE ×7</span>
<span class='wpill wl'>PB ×7</span><span class='wpill wl'>Top3 ×6</span>
<span class='wpill wl'>Top5 ×6</span><span class='wpill wl'>Top10 ×6</span>
<span class='wpill wl'>Top20 ×6</span><span class='wpill wl'>StDev ×9</span>
</div>""", unsafe_allow_html=True)

# ── Hero ───────────────────────────────────────────────────────────────────────
st.markdown("""<div class='hero'>
<div class='hero-badge'>NGEN Markets</div>
<div class='hero-title'>📊 Equity Fund Rankings Generator</div>
<div class='hero-sub'>
  2,392 Equity Scheme funds + 172 Hybrid Equity Savings = <strong style='color:#fff'>2,564 funds total</strong>
  &nbsp;·&nbsp; All plan types &nbsp;·&nbsp; GPT extracts 13 metrics &nbsp;·&nbsp; Scores & ranks &nbsp;·&nbsp; Export CSV
</div>
</div>""", unsafe_allow_html=True)

# ── Step 1: Upload ─────────────────────────────────────────────────────────────
st.markdown("<div class='section-hdr'>1 · Upload SchemeData CSV</div>",
            unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Upload CSV", type=["csv"],
    help="SchemeData2301262313SS.csv",
    label_visibility="collapsed",
)

if uploaded:
    funds_df, total_csv = load_funds(uploaded)
    funds_df["Plan Type"] = funds_df["Scheme NAV Name"].apply(get_plan_type)
    st.session_state.funds_df   = funds_df
    st.session_state.total_csv  = total_csv

if st.session_state.funds_df is not None:
    funds_df  = st.session_state.funds_df
    total_csv = st.session_state.total_csv
    total     = len(funds_df)
    batches   = math.ceil(total / batch_size)

    # Split equity vs hybrid for display
    eq_count  = funds_df["Scheme Category"].str.startswith("Equity Scheme").sum()
    hyb_count = (funds_df["Scheme Category"] == "Hybrid Scheme - Equity Savings").sum()

    # Category pills
    cat_counts = funds_df["Scheme Category"].value_counts()
    cat_pills  = ""
    for cat, cnt in cat_counts.items():
        is_hybrid = cat == "Hybrid Scheme - Equity Savings"
        cls  = "cat-pill hybrid" if is_hybrid else "cat-pill"
        label = cat.replace("Equity Scheme - ", "").replace("Hybrid Scheme - ", "")
        cat_pills += f'<span class="{cls}">{label}<span>{cnt}</span></span>'

    # Plan type pills
    plan_css = {
        "Regular - Growth": "plan-rg", "Direct - Growth": "plan-dg",
        "Regular - IDCW":   "plan-ri", "Direct - IDCW":   "plan-di",
        "Regular - Other":  "plan-ot", "Direct - Other":  "plan-ot",
    }
    plan_counts = funds_df["Plan Type"].value_counts()
    plan_pills  = "".join(
        f'<span class="plan-pill {plan_css.get(p,"plan-ot")}">{p}: {n}</span>'
        for p, n in plan_counts.items()
    )

    st.markdown(f"""<div class='filter-box'>
✅ <strong>All 2,564 equity-related funds loaded automatically</strong><br><br>
Total rows in CSV: <strong>{total_csv:,}</strong> &nbsp;→&nbsp;
Equity Scheme: <strong>{eq_count:,}</strong> &nbsp;+&nbsp;
Hybrid Equity Savings: <strong style='color:#c4b5fd'>{hyb_count:,}</strong> &nbsp;=&nbsp;
<strong style='color:#ffffff;font-size:1rem'>{total:,} funds</strong><br><br>
<strong>Categories:</strong><div class='cat-grid'>{cat_pills}</div>
<strong>Plan types:</strong><div class='plan-grid'>{plan_pills}</div>
</div>""", unsafe_allow_html=True)

    st.markdown(f"""<div class='metric-row'>
{mcard("Total Funds",    f"{total:,}",                         "2,392 equity + 172 hybrid")}
{mcard("Equity Scheme",  f"{eq_count:,}",                      "12 sub-categories")}
{mcard("Hybrid Equity",  f"{hyb_count:,}",                     "Equity Savings")}
{mcard("Unique Schemes", f"{funds_df['Scheme Name'].nunique():,}", "underlying portfolios")}
{mcard("API Batches",    f"{batches}",                          f"{batch_size} funds/call")}
{mcard("Est. Time",      f"~{math.ceil(batches*11/60)} min",   "at ~11 sec/batch")}
{mcard("Est. Cost",      f"~${round(total*0.000072,2)}",        "gpt-4o-mini")}
</div>""", unsafe_allow_html=True)

    st.markdown("""<div class='warn-box'>
⚠️ <strong>Note on plan variants:</strong> Each underlying scheme appears 4–5 times
(Regular Growth, Direct Growth, Regular IDCW, Direct IDCW, etc.).
Portfolio metrics like PE, PB, Sharpe, Holdings % are <strong>identical</strong> across all variants —
only <strong>TER</strong> and <strong>AUM</strong> differ. GPT is instructed to handle this correctly.
</div>""", unsafe_allow_html=True)

    with st.expander(f"👁  Preview all {total:,} funds", expanded=False):
        st.dataframe(
            funds_df[["Scheme Category","Plan Type","AMC","Scheme NAV Name","Launch Date"]],
            use_container_width=True, height=340,
        )

# ── Step 2: Run ────────────────────────────────────────────────────────────────
st.markdown("<div class='section-hdr'>2 · Extract Metrics & Generate Rankings</div>",
            unsafe_allow_html=True)

c1, c2, c3 = st.columns([2, 1, 1])
can_run = (bool(api_key) and
           st.session_state.funds_df is not None and
           not st.session_state.running)

with c1: run_clicked   = st.button("🚀 Start Extraction & Ranking",
                                    disabled=not can_run, use_container_width=True)
with c2: stop_clicked  = st.button("⏹ Stop",  disabled=not st.session_state.running,
                                    use_container_width=True)
with c3: reset_clicked = st.button("🔄 Reset", disabled=st.session_state.running,
                                    use_container_width=True)

if stop_clicked:
    st.session_state.running = False
    st.warning("Stopped. Partial results shown below.")

if reset_clicked:
    for k in ["results","logs","processed","errors","total","final_df","start_time"]:
        st.session_state[k] = ([] if k in ("results","logs")
                                else None if k in ("final_df","start_time")
                                else 0)
    st.rerun()

prog_ph = st.empty()
log_ph  = st.empty()

def render_progress():
    done   = st.session_state.processed
    total  = st.session_state.total or 1
    pct    = done / total
    elapsed = ""
    if st.session_state.start_time:
        secs = int(time.time() - st.session_state.start_time)
        elapsed = f"{secs//60}m {secs%60}s"
        if 0 < done < total:
            eta = int((total - done) * secs / done)
            elapsed += f"  ·  ETA {eta//60}m {eta%60}s"
    with prog_ph.container():
        st.markdown(
            f'<div class="progress-card">'
            f'<div style="display:flex;justify-content:space-between;'
            f'font-size:.85rem;margin-bottom:.5rem;">'
            f'<span><span class="badge-run">{done:,}</span> / {total:,} funds processed</span>'
            f'<span style="color:#4b6080">{elapsed}</span>'
            f'<span><span class="badge-err">{st.session_state.errors}</span> batch errors</span>'
            f'<span style="color:#60a5fa;font-weight:600">{pct*100:.1f}%</span>'
            f'</div></div>', unsafe_allow_html=True)
        st.progress(pct)
    if st.session_state.logs:
        log_ph.markdown(
            f'<div class="log-box">'
            f'{chr(10).join(st.session_state.logs[-80:])}'
            f'</div>', unsafe_allow_html=True)

if st.session_state.processed > 0 or st.session_state.running:
    render_progress()

# ── Extraction loop ────────────────────────────────────────────────────────────
if run_clicked and st.session_state.funds_df is not None:
    if not api_key:
        st.error("Enter your OpenAI API key in the sidebar.")
        st.stop()

    client     = OpenAI(api_key=api_key)
    funds_df   = st.session_state.funds_df
    fund_names = funds_df["Scheme NAV Name"].tolist()
    total      = len(fund_names)
    batches    = math.ceil(total / batch_size)
    ts         = lambda: datetime.now().strftime("%H:%M:%S")

    st.session_state.update({
        "running": True, "results": [], "logs": [],
        "processed": 0, "errors": 0, "total": total,
        "start_time": time.time(), "final_df": None,
    })

    eq_cnt  = funds_df["Scheme Category"].str.startswith("Equity Scheme").sum()
    hyb_cnt = (funds_df["Scheme Category"] == "Hybrid Scheme - Equity Savings").sum()

    # Apply max batches limit
    effective_batches = min(batches, max_batches) if max_batches > 0 else batches
    effective_total   = min(total, effective_batches * batch_size)

    st.session_state.logs.append(
        f"[{ts()}] ▶ Starting  —  {effective_total:,} funds  ·  {effective_batches} batches  ·  model={model}")
    st.session_state.logs.append(
        f"[{ts()}]   Equity Scheme: {eq_cnt:,}  ·  Hybrid Equity Savings: {hyb_cnt:,}")
    st.session_state.logs.append(
        f"[{ts()}]   Batch size: {batch_size}  ·  Delay: {delay}s  ·  Retries: 3"
        + (f"  ·  Max batches: {max_batches}" if max_batches > 0 else "  ·  All batches"))

    for idx in range(effective_batches):
        if not st.session_state.running:
            break

        s     = idx * batch_size
        e     = min(s + batch_size, total)
        batch = fund_names[s:e]

        st.session_state.logs.append(
            f"[{ts()}] 📦 Batch {idx+1:>3}/{effective_batches}  "
            f"[{s+1:>4} – {e:>4}]  ({e-s} funds)")

        records = call_openai(client, batch, model)

        if not records:
            st.session_state.errors += 1
            st.session_state.logs.append(
                f"[{ts()}]   ⚠️  Failed / empty response — filling nulls")
            records = [{"Fund Name": n} for n in batch]
        else:
            st.session_state.logs.append(
                f"[{ts()}]   ✅ {len(records)} records received")

        # Always pin fund name to our source — never trust GPT's version
        for i, rec in enumerate(records):
            if i < len(batch):
                rec["Fund Name"] = batch[i]
            st.session_state.results.append(rec)

        st.session_state.processed = e
        render_progress()

        if idx < effective_batches - 1 and st.session_state.running:
            time.sleep(delay)

    # ── Build scores & ranks ───────────────────────────────────────────────────
    st.session_state.logs.append(f"[{ts()}] ⚙️  Calculating scores and ranking…")

    results_df = pd.DataFrame(st.session_state.results)

    # Ensure all metric columns exist and are numeric
    for col in COLUMNS[1:]:
        if col not in results_df.columns:
            results_df[col] = None
        results_df[col] = results_df[col].apply(clean_num)

    # Score
    results_df["Score"] = results_df.apply(
        lambda r: calculate_score({c: r.get(c) for c in COLUMNS[1:]}),
        axis=1,
    )

    # Sort descending, assign rank
    results_df = results_df.sort_values(
        "Score", ascending=False, na_position="last"
    ).reset_index(drop=True)
    results_df["Rank"] = range(1, len(results_df) + 1)

    # Final column order — matches CSV exactly: metrics + Score + Rank (no extra metadata)
    final_cols = COLUMNS + ["Score", "Rank"]
    for c in final_cols:
        if c not in results_df.columns:
            results_df[c] = None
    results_df = results_df[final_cols]

    st.session_state.final_df = results_df
    st.session_state.running  = False
    st.session_state.logs.append(
        f"[{ts()}] 🏆 Complete — {len(results_df):,} funds ranked!")
    render_progress()
    st.success(f"✅ Done! {len(results_df):,} funds ranked.")
    st.rerun()

# ── Step 3: Results ────────────────────────────────────────────────────────────
st.markdown("<div class='section-hdr'>3 · Results & Export</div>",
            unsafe_allow_html=True)

if st.session_state.final_df is not None:
    df       = st.session_state.final_df
    scored   = df["Score"].notna().sum()
    top_fund = str(df.iloc[0]["Fund Name"]) if len(df) else "—"

    st.markdown(f"""<div class='metric-row'>
{mcard("Total Ranked",    f"{len(df):,}",  "all plan types")}
{mcard("With Score",      f"{scored:,}",   f"{scored/max(len(df),1)*100:.1f}% coverage")}
{mcard("🥇 Rank 1", (top_fund[:35]+"…") if len(top_fund)>35 else top_fund)}
</div>""", unsafe_allow_html=True)

    st.caption(f"Showing {min(100, len(df)):,} of {len(df):,} funds")

    fmt = {
        "AUM Cr.":"{:,.0f}", "TER":"{:.2f}", "PE":"{:.2f}", "PB":"{:.2f}",
        "Top 3 Holdings":"{:.2f}","Top 5 Holdings":"{:.2f}",
        "Top 10 Holdings":"{:.2f}","Top 20 Holdings":"{:.2f}",
        "Sharpe":"{:.3f}","Sortino":"{:.3f}",
        "St Dev":"{:.2f}","Inception":"{:.2f}","Age in Yrs":"{:.1f}",
        "Score":"{:,.4f}","Rank":"{:.0f}",
    }
    st.dataframe(
        df.head(100).style.format(fmt, na_rep="—"),
        use_container_width=True, height=520,
    )

    # Download — Excel format matching sample
    excel_bytes = build_excel_export(df)
    st.download_button(
        "⬇️  Download Rankings_Master.xlsx",
        excel_bytes, "Rankings_Master.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

else:
    st.markdown("""<div class='info-box'>
📌 Upload the SchemeData CSV above → enter your OpenAI key in the sidebar
→ click <strong>🚀 Start Extraction & Ranking</strong>.<br><br>
The app will automatically process <strong>all 2,564 funds</strong>:
<strong>2,392</strong> Equity Scheme (12 categories) +
<strong>172</strong> Hybrid Scheme - Equity Savings —
across all plan types (Regular/Direct · Growth/IDCW/Other).<br><br>
<strong>~52 batches · ~10 minutes · ~$0.18</strong> on gpt-4o-mini.
</div>""", unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("""
<hr style='border-color:#1e2a45;margin-top:3rem;'>
<div style='text-align:center;color:#2d3f5c;font-size:.78rem;padding-bottom:1rem;'>
  NGEN Markets · Equity Fund Rankings Engine ·
  2,392 Equity Scheme + 172 Hybrid Equity Savings · Powered by OpenAI GPT
</div>""", unsafe_allow_html=True)
